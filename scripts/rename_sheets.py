import os
import re
from openpyxl import load_workbook

# Configurações
pasta = './sheets'  # Pasta onde os arquivos estão localizados
extensoes = ('.xlsx', '.xlsm')  # Extensões suportadas

# Função para sanitizar nomes de arquivo
def sanitizar_nome(nome):
    return re.sub(r'[\\/*?:"<>|]', '_', nome)

# Processar cada arquivo na pasta
for arquivo in os.listdir(pasta):
    if arquivo.endswith(extensoes):
        caminho_antigo = os.path.join(pasta, arquivo)
        
        try:
            wb = load_workbook(caminho_antigo, read_only=True)
            sheet = wb.active  # Pega a planilha ativa
            
            # Lê o valor da célula A2 (linha 2, coluna 1)
            novo_nome_cel = sheet.cell(row=2, column=1).value
            
            if novo_nome_cel:
                # Remove espaços extras e sanitiza
                novo_nome = sanitizar_nome(str(novo_nome_cel).strip())
                extensao = os.path.splitext(arquivo)[1]  # Mantém a extensão original
                caminho_novo = os.path.join(pasta, f"{novo_nome}{extensao}")
                
                # Evita sobrescrever arquivos existentes
                contador = 1
                while os.path.exists(caminho_novo):
                    caminho_novo = os.path.join(pasta, f"{novo_nome}_{contador}{extensao}")
                    contador += 1
                
                # Renomeia o arquivo
                os.rename(caminho_antigo, caminho_novo)
                print(f"✅ Renomeado: {arquivo} -> {novo_nome}{extensao}")
            else:
                print(f"⛔ Célula A2 vazia em: {arquivo} (ignorado)")
                
        except Exception as e:
            print(f"⚠️ Erro ao processar {arquivo}: {str(e)}")
        finally:
            wb.close()  # Fecha o workbook